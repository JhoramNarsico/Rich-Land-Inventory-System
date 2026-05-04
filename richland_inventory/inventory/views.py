
import csv
import json
import uuid
from datetime import timedelta, datetime
from decimal import Decimal, InvalidOperation

# --- THIRD-PARTY IMPORTS ---
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView
from django.contrib.messages.views import SuccessMessageMixin
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from django.db import models, transaction
from django.db.models import Q, F, Sum, Count, ExpressionWrapper, DecimalField, Value, OuterRef, Subquery
from django.db.models.functions import TruncDate, Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, TemplateView
from django.views.generic.edit import CreateView, UpdateView, DeleteView

from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiTypes
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, permissions, filters

# --- LOCAL IMPORTS ---
from core.cache_utils import clear_dashboard_cache
from . import importers as inventory_importers
from .exports import (
    generate_sow_history_export, generate_expense_report, generate_customer_list_export,
    generate_customer_statement, generate_supplier_deliveries_export
)
from .forms import (
    AnalyticsFilterForm, CategoryCreateForm, CustomerFilterForm, CustomerForm, 
    CustomerPaymentForm, ExpenseFilterForm, ExpenseForm, ProductCreateForm, 
    ProductFilterForm, ProductHistoryFilterForm, ProductUpdateForm, 
    PurchaseOrderFilterForm, RefundForm, StockOutForm, StockTransactionForm, 
    TransactionFilterForm, TransactionReportForm
)
from .models import (
    Category, Customer, CustomerPayment, Expense, ExpenseCategory, HydraulicSow,
    POSSale, PriceOverrideLog, Product, PurchaseOrder, PurchaseOrderItem,
    StockTransaction, Supplier
)
from .serializers import (
    ProductSerializer, CategorySerializer, CustomerSerializer, CustomerPaymentSerializer,
    HydraulicSowSerializer, POSSaleSerializer, ExpenseSerializer, ExpenseCategorySerializer
)
from .utils import render_to_pdf
from .utils_service import get_service_product



from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from .models import POSSale, Product, StockTransaction, Category
from decimal import Decimal

# --- MISC ---

@login_required
def feedback_view(request):
    return render(request, 'inventory/feedback.html', {'page_title': 'User Feedback'})

# --- REFUND PORTAL ---

@login_required
@permission_required('inventory.can_adjust_stock', raise_exception=True)
def refund_portal(request):
    return render(request, 'inventory/refund.html')

@login_required
@permission_required('inventory.can_adjust_stock', raise_exception=True)
def refund_search(request):
    rid = request.GET.get('rid')
    sale = POSSale.objects.filter(receipt_id=rid).first()
    if not sale:
        return JsonResponse({'status': 'error', 'message': 'Receipt not found'})
    
    items = []
    # Fetch sold items and calculate remaining returnable quantity
    for item in sale.items.filter(transaction_type='OUT', transaction_reason='SALE'):
        # Skip Hydraulic Service jobs
        if "Hydraulic Service" in item.product.name or item.product.sku == "SVC-HYD-001":
            continue
            
        returned = sale.items.filter(
            product=item.product, 
            transaction_type='IN', 
            transaction_reason='RETURN'
        ).aggregate(qty=Sum('quantity'))['qty'] or 0
        
        remaining = item.quantity - returned
        if remaining > 0:
            items.append({'id': item.id, 'name': item.product.name, 'qty': remaining})
            
    return JsonResponse({'status': 'success', 'items': items})

@login_required
@permission_required('inventory.can_adjust_stock', raise_exception=True)
@transaction.atomic
def refund_process(request):
    if request.method == 'POST':
        rid = request.POST.get('receipt_id')
        reason = request.POST.get('refund_reason', '').strip()
        
        if not reason:
            messages.error(request, "A reason for the refund is required.")
            return redirect('inventory:refund_portal')

        sale = POSSale.objects.get(receipt_id=rid)
        
        # Check if the receipt belongs to a Hydraulic SOW job
        if sale.notes and "Hydraulic Job" in sale.notes:
            messages.error(request, "Refunds for Hydraulic Service Jobs are not permitted.")
            return redirect('inventory:refund_portal')
        
        refund_count = 0
        for key, qty in request.POST.items():
            if key.startswith('qty_') and int(qty) > 0:
                item_id = key.split('_')[1]
                item = sale.items.get(id=item_id)
                
                # Perform the return logic
                StockTransaction.objects.create(
                    product=item.product,
                    pos_sale=sale,
                    transaction_type='IN',
                    transaction_reason='RETURN',
                    quantity=int(qty),
                    selling_price=item.selling_price,
                    user=request.user,
                    notes=f"Refund for {rid}. Reason: {reason}"
                )
                item.product.quantity += int(qty)
                item.product.save()
                refund_count += 1
        
        if refund_count == 0:
            messages.warning(request, "No items were selected for refund.")
        else:
            messages.success(request, f"Refunds processed for receipt {rid}.")
        
        return redirect('inventory:refund_portal')
    return redirect('inventory:refund_portal')

@login_required
@permission_required('inventory.add_hydraulicsow', raise_exception=True)
def hydraulic_sow_create(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    next_url = request.GET.get('next')

    if request.method == 'POST':
        cost_input = request.POST.get('cost')
        cost_decimal = Decimal('0.00')
        if cost_input:
            try:
                cost_decimal = Decimal(str(cost_input))
            except (ValueError, TypeError, InvalidOperation):
                pass

        charge_account = request.POST.get('charge_account')
        mark_paid = request.POST.get('mark_paid')

        # --- VALIDATION: Required Fields ---
        hose_type = request.POST.get('hose_type', '').strip()
        diameter = request.POST.get('diameter', '').strip()
        length = request.POST.get('length', '').strip()
        fitting_a = request.POST.get('fitting_a', '').strip()
        fitting_b = request.POST.get('fitting_b', '').strip()

        if not all([hose_type, diameter, length, fitting_a, fitting_b]):
            messages.error(request, "Please fill in all required fields: Hose Type, Diameter, Length, and Fittings.")
            sow_data = HydraulicSow(
                customer=customer, hose_type=hose_type,
                diameter=diameter, length=request.POST.get('length') or None,
                pressure=request.POST.get('pressure') or None, application=request.POST.get('application', ''),
                fitting_a=fitting_a, fitting_b=fitting_b,
                orientation=request.POST.get('orientation') or None, protection=request.POST.get('protection', ''),
                cost=cost_decimal if cost_decimal > 0 else None, notes=request.POST.get('notes', '')
            )
            return render(request, 'inventory/hydraulic_sow_form.html', {
                'customer': customer, 'sow': sow_data, 'page_title': 'Create Hydraulic SOW', 
                'is_charged': False, 'next_url': next_url, 'is_charge_checked': charge_account, 'is_paid_checked': mark_paid})

        # --- VALIDATION for named customers ---
        if customer.name != "Walk-in Customer":
            has_cost = cost_decimal > 0
            has_payment_method = charge_account or mark_paid
            error_message = None

            if not has_cost:
                error_message = "A service cost is required for all customer jobs."
            elif not has_payment_method:
                error_message = "You must select a payment method (Charge to Account or Pay Cash)."

            if error_message:
                messages.error(request, error_message)
                sow_data = HydraulicSow(
                    customer=customer, hose_type=request.POST.get('hose_type', ''),
                    diameter=request.POST.get('diameter', ''), length=request.POST.get('length') or None,
                    pressure=request.POST.get('pressure') or None, application=request.POST.get('application', ''),
                    fitting_a=request.POST.get('fitting_a', ''), fitting_b=request.POST.get('fitting_b', ''),
                    orientation=request.POST.get('orientation') or None, protection=request.POST.get('protection', ''),
                    cost=cost_decimal if cost_decimal > 0 else None, notes=request.POST.get('notes', '')
                )
                return render(request, 'inventory/hydraulic_sow_form.html', {
                    'customer': customer, 
                    'sow': sow_data, 
                    'page_title': 'Create Hydraulic SOW', 
                    'is_charged': False, 
                    'next_url': next_url,
                    'is_charge_checked': charge_account,
                    'is_paid_checked': mark_paid})
        
        # --- VALIDATION for Walk-in Customers ---
        elif customer.name == "Walk-in Customer" and cost_decimal <= 0:
            messages.error(request, "A service cost is required for Walk-in Customers.")
            sow_data = HydraulicSow(
                customer=customer, hose_type=request.POST.get('hose_type', ''),
                diameter=request.POST.get('diameter', ''), length=request.POST.get('length') or None,
                pressure=request.POST.get('pressure') or None, application=request.POST.get('application', ''),
                cost=None, notes=request.POST.get('notes', '')
            )
            return render(request, 'inventory/hydraulic_sow_form.html', {'customer': customer, 'sow': sow_data, 'page_title': 'Create Hydraulic SOW', 'is_charged': False, 'next_url': next_url})

        with transaction.atomic():
            sow = HydraulicSow.objects.create(
                customer=customer,
                created_by=request.user,
                hose_type=request.POST.get('hose_type', ''),
                diameter=request.POST.get('diameter', ''),
                length=request.POST.get('length') or None,
                pressure=request.POST.get('pressure') or None,
                application=request.POST.get('application', ''),
                fitting_a=request.POST.get('fitting_a', ''),
                fitting_b=request.POST.get('fitting_b', ''),
                orientation=request.POST.get('orientation') or None,
                protection=request.POST.get('protection', ''),
                cost=cost_decimal if cost_decimal > 0 else None,
                notes=request.POST.get('notes', '')
            )

            if cost_decimal > 0:
                # For Walk-in Customers, always generate a receipt (Cash default) if there is a cost.
                # For named customers, only generate if charged (Credit) or marked paid (Cash).
                if charge_account or mark_paid or customer.name == "Walk-in Customer":
                    payment_method = 'CREDIT' if charge_account else 'CASH'
                    amount_paid = 0 if charge_account else cost_decimal

                    receipt_id = sow.sow_id
                    sale_record = POSSale.objects.create(
                        receipt_id=receipt_id,
                        customer=customer,
                        cashier=request.user,
                        payment_method=payment_method,
                        total_amount=cost_decimal,
                        amount_paid=amount_paid,
                        change_given=0,
                        notes=f"Hydraulic Job #{sow.id}: {sow.hose_type} ({sow.application})"
                    )

                    # LOG INDIVIDUAL ITEM IN TRANSACTION LOG
                    service_product = get_service_product()
                    StockTransaction.objects.create(
                        product=service_product,
                        pos_sale=sale_record,
                        transaction_type='OUT',
                        transaction_reason=StockTransaction.TransactionReason.SALE,
                        quantity=1,
                        selling_price=cost_decimal,
                        user=request.user,
                        notes=f"Hydraulic Service: {sow.hose_type} | {sow.diameter}\" | {sow.application}"
                    )

                    messages.success(request, f"Hydraulic SOW saved. Receipt and transaction log generated.")
                    return redirect('inventory:pos_receipt_detail', receipt_id=receipt_id)

        messages.success(request, f"Hydraulic Scope of Work saved for {customer.name}")
            
        if next_url:
            return redirect(next_url)
            
        return redirect('inventory:customer_detail', pk=pk)

    return render(request, 'inventory/hydraulic_sow_form.html', {
        'customer': customer,
        'page_title': 'Create Hydraulic SOW',
        'is_charged': False,
        'next_url': next_url,
    })

@login_required
@permission_required('inventory.change_hydraulicsow', raise_exception=True)
def hydraulic_sow_update(request, pk, sow_pk):
    customer = get_object_or_404(Customer, pk=pk)
    sow = get_object_or_404(HydraulicSow, pk=sow_pk, customer=customer)
    
    # Ensure SOW ID exists (for legacy records)
    if not sow.sow_id:
        sow.save()

    if request.method == 'POST':
        cost_input = request.POST.get('cost')
        cost_decimal = Decimal('0.00')
        if cost_input:
            try:
                cost_decimal = Decimal(str(cost_input))
            except (ValueError, TypeError, InvalidOperation):
                pass

        charge_to_account = request.POST.get('charge_account')
        mark_paid = request.POST.get('mark_paid')
        ledger_entry = POSSale.objects.filter(receipt_id=sow.sow_id).first()
        if not ledger_entry:
            ledger_entry = POSSale.objects.filter(receipt_id=f"SOW-{sow.id}").first()

        # --- VALIDATION: Required Fields ---
        hose_type = request.POST.get('hose_type', '').strip()
        diameter = request.POST.get('diameter', '').strip()
        length = request.POST.get('length', '').strip()
        fitting_a = request.POST.get('fitting_a', '').strip()
        fitting_b = request.POST.get('fitting_b', '').strip()

        if not all([hose_type, diameter, length, fitting_a, fitting_b]):
            messages.error(request, "Please fill in all required fields: Hose Type, Diameter, Length, and Fittings.")
            sow.hose_type = hose_type
            sow.diameter = diameter
            sow.fitting_a = fitting_a
            sow.fitting_b = fitting_b
            sow.length = request.POST.get('length') or None
            sow.pressure = request.POST.get('pressure') or None
            sow.application = request.POST.get('application', '')
            sow.orientation = request.POST.get('orientation') or None
            sow.protection = request.POST.get('protection', '')
            sow.notes = request.POST.get('notes', '')
            sow.cost = cost_decimal if cost_decimal > 0 else None
            return render(request, 'inventory/hydraulic_sow_form.html', {'customer': customer, 'sow': sow, 'page_title': f'Edit Hydraulic SOW {sow.sow_id or sow.id}', 'is_charged': ledger_entry is not None, 'is_charge_checked': charge_to_account, 'is_paid_checked': mark_paid})

        # --- VALIDATION for named customers on un-charged SOWs ---
        if not ledger_entry and customer.name != "Walk-in Customer":
            has_cost = cost_decimal > 0
            has_payment_method = charge_to_account or mark_paid
            error_message = None

            if not has_cost:
                error_message = "A service cost is required to create a new charge for this job."
            elif not has_payment_method:
                error_message = "You must select a payment method (Charge or Pay Cash) to create a new charge."
            
            if error_message:
                messages.error(request, error_message)
                sow.hose_type = request.POST.get('hose_type', ''); sow.diameter = request.POST.get('diameter', '')
                sow.length = request.POST.get('length') or None; sow.pressure = request.POST.get('pressure') or None
                sow.application = request.POST.get('application', ''); sow.fitting_a = request.POST.get('fitting_a', '')
                sow.fitting_b = request.POST.get('fitting_b', ''); sow.orientation = request.POST.get('orientation') or None
                sow.protection = request.POST.get('protection', ''); sow.notes = request.POST.get('notes', '')
                sow.cost = cost_decimal if cost_decimal > 0 else None
                return render(request, 'inventory/hydraulic_sow_form.html', {
                    'customer': customer, 'sow': sow, 
                    'page_title': f'Edit Hydraulic SOW {sow.sow_id or sow.id}', 
                    'is_charged': False,
                    'is_charge_checked': charge_to_account,
                    'is_paid_checked': mark_paid
                })
        
        # --- VALIDATION for Walk-in Customers ---
        elif not ledger_entry and customer.name == "Walk-in Customer" and cost_decimal <= 0:
            messages.error(request, "A service cost is required for Walk-in Customers.")
            sow.hose_type = request.POST.get('hose_type', ''); sow.diameter = request.POST.get('diameter', '')
            sow.cost = None
            return render(request, 'inventory/hydraulic_sow_form.html', {
                'customer': customer, 'sow': sow, 
                'page_title': f'Edit Hydraulic SOW {sow.sow_id or sow.id}', 'is_charged': False
            })

        # Update SOW fields
        sow.hose_type = request.POST.get('hose_type', '')
        sow.diameter = request.POST.get('diameter', '')
        sow.length = request.POST.get('length') or None
        sow.pressure = request.POST.get('pressure') or None
        sow.application = request.POST.get('application', '')
        sow.fitting_a = request.POST.get('fitting_a', '')
        sow.fitting_b = request.POST.get('fitting_b', '')
        sow.orientation = request.POST.get('orientation') or None
        sow.protection = request.POST.get('protection', '')
        sow.notes = request.POST.get('notes', '')

        sow.cost = cost_decimal if cost_decimal > 0 else None
        sow.save()

        # Handle charging logic
        if ledger_entry:
            if ledger_entry.total_amount != cost_decimal:
                ledger_entry.total_amount = cost_decimal
                ledger_entry.save()

                # Update existing transaction if it exists
                st = StockTransaction.objects.filter(pos_sale=ledger_entry).first()
                if st:
                    st.selling_price = cost_decimal
                    st.notes = f"Hydraulic Service: {sow.hose_type} | {sow.diameter}\" | {sow.application} (Updated)"
                    st.save()

                messages.success(request, f"SOW updated. Associated charge was adjusted to ₱{cost_decimal:,.2f}.")
            else:
                messages.success(request, "SOW updated. No changes to the associated charge.")
        elif (charge_to_account or mark_paid or customer.name == "Walk-in Customer") and cost_decimal > 0:
            payment_method = 'CREDIT' if charge_to_account else 'CASH'
            amount_paid = 0 if charge_to_account else cost_decimal
            
            with transaction.atomic():
                sale_record = POSSale.objects.create(
                    receipt_id=sow.sow_id, 
                    customer=customer, 
                    cashier=request.user, 
                    payment_method=payment_method, 
                    total_amount=cost_decimal, 
                    amount_paid=amount_paid,
                    notes=f"Hydraulic Job #{sow.id}: {sow.hose_type} ({sow.application})"
                )

                # LOG INDIVIDUAL ITEM IN TRANSACTION LOG
                service_product = get_service_product()
                StockTransaction.objects.create(
                    product=service_product,
                    pos_sale=sale_record,
                    transaction_type='OUT',
                    transaction_reason=StockTransaction.TransactionReason.SALE,
                    quantity=1,
                    selling_price=cost_decimal,
                    user=request.user,
                    notes=f"Hydraulic Service: {sow.hose_type} | {sow.diameter}\" | {sow.application}"
                )

            messages.success(request, f"SOW updated and a new charge of ₱{cost_decimal:,.2f} was added to the account.")
        else:
            messages.success(request, "Hydraulic SOW updated successfully.")
        return redirect('inventory:customer_detail', pk=pk)

    ledger_entry = POSSale.objects.filter(receipt_id=sow.sow_id).first()
    if not ledger_entry:
        ledger_entry = POSSale.objects.filter(receipt_id=f"SOW-{sow.id}").first()
    
    # Attach ledger to sow object temporarily for template logic
    sow.pos_sale = ledger_entry
    
    return render(request, 'inventory/hydraulic_sow_form.html', {'customer': customer, 'sow': sow, 'page_title': f'Edit Hydraulic SOW {sow.sow_id or sow.id}', 'is_charged': ledger_entry is not None})

@login_required
@permission_required('inventory.add_hydraulicsow', raise_exception=True)
def hydraulic_sow_import(request):
    if request.method == 'POST':
        # Handle file upload and parsing logic here
        messages.info(request, "Import functionality is under construction.")
        return redirect('inventory:customer_list')
        
    # You will need a simple template for this, or reuse a generic import template
    return render(request, 'inventory/form_import.html', {
        'title': 'Import Hydraulic SOW'
    })

@login_required
@permission_required('inventory.view_hydraulicsow', raise_exception=True)
def export_sow_history(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    format_type = request.GET.get('format', 'pdf')
    sow_q = request.GET.get('sow_q', '')

    sows = customer.sows.select_related('created_by').all()
    
    if sow_q:
        q_sow = Q(sow_id__icontains=sow_q) | \
                Q(hose_type__icontains=sow_q) | \
                Q(application__icontains=sow_q) | \
                Q(notes__icontains=sow_q) | \
                Q(fitting_a__icontains=sow_q) | \
                Q(fitting_b__icontains=sow_q)
        if sow_q.isdigit():
            q_sow |= Q(id=sow_q)
        sows = sows.filter(q_sow)

    response = generate_sow_history_export(customer, sows, format_type, request)
    if response:
        return response
    return HttpResponse("Error Generating Export", status=500)

@login_required
@permission_required('inventory.add_hydraulicsow', raise_exception=True)
def import_sow_history(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == "POST" and request.FILES.get('csv_file'):
        file_obj = request.FILES['csv_file']
        if not (file_obj.name.lower().endswith('.csv') or file_obj.name.lower().endswith('.xlsx')):
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect('inventory:customer_detail', pk=pk)

        try:
            count, errors = inventory_importers.import_sow_from_file(file_obj, customer, request.user)
            
            if errors:
                for error in errors[:5]:
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.warning(request, f"And {len(errors) - 5} more errors...")
            else:
                if count > 0:
                    messages.success(request, f"Successfully imported {count} SOW records.")
                else:
                    messages.info(request, "Import complete. No new SOW records were added.")

        except Exception as e:
            messages.error(request, f"An unexpected error occurred while processing the file: {e}")
            
        return redirect('inventory:customer_detail', pk=pk)

    instructions = {
        "title": "How to Format Your SOW Data",
        "general":[
            "Fill in your SOW (Scope of Work) data in the downloaded Excel template.",
            "Do NOT change the column headers in the 'Data' sheet.",
            "Fields marked with an asterisk (*) are REQUIRED.",
            "Delete the sample row in the 'Data' sheet before uploading."
        ],
        "columns":[
            {"name": "Hose Type (*)", "desc": "The type of hose.", "example": "'2 Wire'"},
            {"name": "Diameter (*)", "desc": "The hose diameter.", "example": "'1/2'"},
            {"name": "Length (*)", "desc": "The length of the hose (as a number).", "example": "1000"},
            {"name": "Pressure", "desc": "The pressure rating (as a number).", "example": "3000"},
            {"name": "Cost", "desc": "The cost of the service (as a number, no currency symbol).", "example": "1500.00"},
            {"name": "Application", "desc": "Where the hose is used.", "example": "'Excavator Boom'"},
            {"name": "Fitting A (*)", "desc": "The type of the first fitting.", "example": "'JIC F'"},
            {"name": "Fitting B (*)", "desc": "The type of the second fitting.", "example": "'BSP M'"},
            {"name": "Notes", "desc": "Any additional notes or comments.", "example": "'Urgent repair'"},
        ]
    }
    return render(request, 'inventory/sow_import.html', {'customer': customer, 'instructions': instructions})

@login_required
def download_sow_template(request):
    """Downloads an Excel template for SOW imports with instructions."""
    wb = Workbook()
    
    # --- Create Instructions Sheet ---
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # Styles
    title_font = Font(name='Calibri', bold=True, size=16, color="1F4E78")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    instruction_header_font = Font(name='Calibri', bold=True, size=12)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Title
    ws_instructions['A1'] = "How to Use This Import Template"
    ws_instructions['A1'].font = title_font
    ws_instructions.merge_cells('A1:D1')

    # General Instructions
    ws_instructions.append([]) # Spacer
    ws_instructions['A3'] = "General Rules"
    ws_instructions['A3'].font = header_font
    ws_instructions.append(["1. Fill in your SOW (Scope of Work) data in the 'Data' sheet."])
    ws_instructions.append(["2. Do NOT change the column headers in the 'Data' sheet."])
    ws_instructions.append(["3. Fields marked with an asterisk (*) are REQUIRED."])
    ws_instructions.append(["4. Delete the sample row in the 'Data' sheet before uploading."])
    ws_instructions.append([]) # Spacer
    
    # Column Descriptions
    ws_instructions['A9'] = "Column Guide"
    ws_instructions['A9'].font = header_font
    
    headers =[
        ("Column", "Description", "Example", "Required?"),
        ("Hose Type", "The type of hose.", "'2 Wire'", "Yes (*)"),
        ("Diameter", "The hose diameter.", "'1/2'", "Yes (*)"),
        ("Length", "The length of the hose (as a number).", "1000", "Yes (*)"),
        ("Pressure", "The pressure rating (as a number).", "3000", "No"),
        ("Cost", "The cost of the service (as a number, no currency symbol).", "1500.00", "No"),
        ("Application", "Where the hose is used.", "'Excavator Boom'", "No"),
        ("Fitting A", "The type of the first fitting.", "'JIC F'", "Yes (*)"),
        ("Fitting B", "The type of the second fitting.", "'BSP M'", "Yes (*)"),
        ("Notes", "Any additional notes or comments.", "'Urgent repair'", "No"),
    ]
    
    for row_data in headers:
        ws_instructions.append(row_data)

    # Styling for instructions table
    for cell in ws_instructions['A10:D10'][0]:
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row in ws_instructions['A10:D19']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col_idx, width in enumerate([20, 50, 25, 15], 1):
        ws_instructions.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Create Data Sheet ---
    ws_data = wb.create_sheet(title="Data")
    data_headers =['Hose Type', 'Diameter', 'Length', 'Pressure', 'Cost', 'Application', 'Fitting A', 'Fitting B', 'Notes']
    ws_data.append(data_headers)
    ws_data.append(['2 Wire', '1/2', 1000, 3000, 1500.00, 'Excavator Boom', 'JIC F', 'BSP M', 'Urgent repair'])
    
    # Style header and set column widths
    for i, cell in enumerate(ws_data['1:1'], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws_data.column_dimensions[get_column_letter(i)].width = 22

    ws_data.freeze_panes = 'A2'
    for row in ws_data['A1:I2']:
        for cell in row:
            cell.border = thin_border

    # --- Prepare Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sow_import_template.xlsx"'
    wb.save(response)
    return response

@login_required
def download_expense_template(request):
    """Downloads an Excel template for expense imports."""
    wb = Workbook()
    
    # --- Create Instructions Sheet ---
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # Styles
    title_font = Font(name='Calibri', bold=True, size=16, color="1F4E78")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws_instructions['A1'] = "How to Use This Import Template"
    ws_instructions['A1'].font = title_font
    ws_instructions.merge_cells('A1:D1')

    # General Instructions
    ws_instructions.append([]) 
    ws_instructions['A3'] = "General Rules"
    ws_instructions['A3'].font = header_font
    ws_instructions.append(["1. Fill in your expense data in the 'Data' sheet."])
    ws_instructions.append(["2. Do NOT change the column headers in the 'Data' sheet."])
    ws_instructions.append(["3. Delete the sample rows before uploading."])
    ws_instructions.append(["4. Use YYYY-MM-DD format for dates."])
    ws_instructions.append([]) 
    
    # Column Descriptions
    ws_instructions['A9'] = "Column Guide"
    ws_instructions['A9'].font = header_font
    
    headers =[
        ("Column", "Description", "Example", "Required?"),
        ("Date", "Expense date (YYYY-MM-DD).", timezone.now().strftime('%Y-%m-%d'), "Yes"),
        ("Category", "Expense Category (e.g. Rent, Utilities).", "'Utilities'", "No"),
        ("Description", "Details about the expense.", "'Office Supplies'", "Yes"),
        ("Amount", "Amount spent.", "150.00", "Yes"),
    ]
    
    for row_data in headers:
        ws_instructions.append(row_data)

    # Styling
    for cell in ws_instructions['A10:D10'][0]:
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row in ws_instructions['A10:D14']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col_idx, width in enumerate([20, 40, 25, 20], 1):
        ws_instructions.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Create Data Sheet ---
    ws_data = wb.create_sheet(title="Data")
    data_headers =['Date', 'Category', 'Description', 'Amount']
    ws_data.append(data_headers)
    
    # Sample Data
    ws_data.append([timezone.now().strftime('%Y-%m-%d'), 'Utilities', 'Electric Bill', 1500.00])
    
    # Style header
    for i, cell in enumerate(ws_data['1:1'], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws_data.column_dimensions[get_column_letter(i)].width = 25

    ws_data.freeze_panes = 'A2'
    for row in ws_data['A1:D2']:
        for cell in row:
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="expense_import_template.xlsx"'
    wb.save(response)
    return response

# --- EXPENSE MANAGEMENT ---

class ExpenseListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Expense
    template_name = 'inventory/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20
    permission_required = 'inventory.view_expense'

    def get_queryset(self):
        queryset = Expense.objects.select_related('category', 'recorded_by').all()
        
        today = timezone.now().date()
        default_month = str(today.month)
        default_year = str(today.year)
        
        data = self.request.GET.copy()
        
        # Default to current month/year if not specified in the URL
        if not data.get('month') and not data.get('year'):
            data['month'] = default_month
            data['year'] = default_year
        
        self.filter_form = ExpenseFilterForm(data)
        if self.filter_form.is_valid():
            if self.filter_form.cleaned_data.get('q'):
                queryset = queryset.filter(description__icontains=self.filter_form.cleaned_data['q'])
            if self.filter_form.cleaned_data.get('category'):
                queryset = queryset.filter(category=self.filter_form.cleaned_data['category'])
            
            m = self.filter_form.cleaned_data.get('month')
            y = self.filter_form.cleaned_data.get('year')
            if y:
                queryset = queryset.filter(expense_date__year=y)
                if m:
                    queryset = queryset.filter(expense_date__month=m)
                
        return queryset.order_by('-expense_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = self.filter_form
        context['total_expenses'] = self.get_queryset().aggregate(total=Sum('amount'))['total'] or 0
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        
        if self.filter_form.is_valid():
            context['current_month'] = self.filter_form.cleaned_data.get('month')
            context['current_year'] = self.filter_form.cleaned_data.get('year')
            
            if context['current_year'] and not context['current_month']:
                context['period_name'] = f"Year {context['current_year']}"
            elif context['current_year'] and context['current_month']:
                try:
                    d = datetime(int(context['current_year']), int(context['current_month']), 1)
                    context['period_name'] = d.strftime('%B %Y')
                except:
                    context['period_name'] = "Selected Period"
            else:
                context['period_name'] = "All Time"
        
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
                
        return context

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_expenses(request, export_format)
        return super().get(request, *args, **kwargs)

    def export_expenses(self, request, format_type):
        expenses = self.get_queryset().order_by('expense_date')
        response = generate_expense_report(expenses, format_type, request)
        if response:
            return response
        messages.error(request, "Could not generate report. Invalid format specified.")
        return redirect('inventory:expense_list')

class ExpenseCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'inventory/expense_form.html'
    success_url = reverse_lazy('inventory:expense_list')
    success_message = "Expense recorded successfully."
    permission_required = 'inventory.add_expense'

    def form_valid(self, form):
        form.instance.recorded_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Record New Expense"
        return context
        
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['target_month'] = self.request.GET.get('month')
        kwargs['target_year'] = self.request.GET.get('year')
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        m = self.request.GET.get('month')
        y = self.request.GET.get('year')
        if y:
            try:
                today = timezone.now().date()
                if m:
                    target_date = datetime(int(y), int(m), 1).date()
                    if target_date.year == today.year and target_date.month == today.month:
                        initial['expense_date'] = today
                    else:
                        initial['expense_date'] = target_date
                else:
                    if str(today.year) == str(y):
                        initial['expense_date'] = today
                    else:
                        initial['expense_date'] = datetime(int(y), 1, 1).date()
            except:
                pass
        return initial

    def get_success_url(self):
        url = reverse_lazy('inventory:expense_list')
        m = self.request.GET.get('month')
        y = self.request.GET.get('year')
        
        params =[]
        if m: params.append(f"month={m}")
        if y: params.append(f"year={y}")
        
        if params:
            return f"{url}?{'&'.join(params)}"
        return url

class ExpenseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'inventory/expense_form.html'
    success_url = reverse_lazy('inventory:expense_list')
    success_message = "Expense updated successfully."
    permission_required = 'inventory.change_expense'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Edit Expense"
        return context

class ExpenseDeleteView(LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, DeleteView):
    model = Expense
    template_name = 'inventory/expense_confirm_delete.html'
    success_url = reverse_lazy('inventory:expense_list')
    success_message = "Expense deleted successfully."
    permission_required = 'inventory.delete_expense'

    def test_func(self):
        expense = self.get_object()
        return self.request.user == expense.recorded_by or self.request.user.is_superuser

@login_required
@permission_required('inventory.add_expense', raise_exception=True)
def import_expenses(request):
    if request.method == "POST" and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        if not (csv_file.name.endswith('.csv') or csv_file.name.endswith('.xlsx')):
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect('inventory:expense_list')

        try:
            data =[]
            if csv_file.name.endswith('.csv'):
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                reader.fieldnames = [name.strip().lower().replace(' ', '_') for name in reader.fieldnames]
                data = list(reader)
            elif csv_file.name.endswith('.xlsx'):
                wb = load_workbook(csv_file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
                    for row in rows[1:]:
                        # Keep native types for date/amount if possible, but handle None
                        row_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
                        data.append(row_dict)
            
            count = 0
            with transaction.atomic():
                for row in data:
                    cat_name = row.get('category')
                    category = None
                    if cat_name:
                        category, _ = ExpenseCategory.objects.get_or_create(name=cat_name.strip())

                    # Handle Date (Excel returns datetime, CSV returns string)
                    date_val = row.get('date')
                    if hasattr(date_val, 'date'): # datetime object
                        date_val = date_val.date()
                    
                    # Handle Amount (Excel returns int/float, CSV returns string)
                    amount_val = row.get('amount', 0)
                    if amount_val is None: amount_val = 0

                    Expense.objects.create(
                        expense_date=date_val,
                        category=category,
                        description=row.get('description', ''),
                        amount=Decimal(str(amount_val)),
                        recorded_by=request.user
                    )
                    count += 1
            messages.success(request, f"Successfully imported {count} expenses.")
        except Exception as e:
            messages.error(request, f"Error processing file: {e}")
            
        return redirect('inventory:expense_list')
        
    return render(request, 'inventory/expense_import.html')

# --- AUTHENTICATION ---

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'

    def form_valid(self, form):
        response = super().form_valid(form)
        user = form.get_user()
        messages.success(self.request, f"Welcome back, {user.username}!")
        return response

# --- CUSTOMER & BILLING MANAGEMENT ---

class CustomerListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Customer
    template_name = 'inventory/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20
    permission_required = 'inventory.view_customer'

    def get_queryset(self):
        # Optimized to avoid Cartesian product issues when summing multiple relations
        credit_sales_subquery = POSSale.objects.filter(
            customer=OuterRef('pk'),
            payment_method='CREDIT'
        ).values('customer').annotate(
            total=Sum('total_amount')
        ).values('total')

        payments_subquery = CustomerPayment.objects.filter(
            customer=OuterRef('pk')
        ).values('customer').annotate(
            total=Sum('amount')
        ).values('total')

        qs = Customer.objects.exclude(name="Walk-in Customer").annotate(
            total_credit_sales=Coalesce(Subquery(credit_sales_subquery), Decimal('0.00'), output_field=DecimalField()),
            total_payments_made=Coalesce(Subquery(payments_subquery), Decimal('0.00'), output_field=DecimalField())
        ).annotate(
            balance=F('total_credit_sales') - F('total_payments_made')
        ).order_by('name')

        status = self.request.GET.get('status')
        if status == 'outstanding':
            qs = qs.filter(balance__gt=0)
        elif status == 'cleared':
            qs = qs.filter(balance__lte=0)

        self.filter_form = CustomerFilterForm(self.request.GET)
        if self.filter_form.is_valid():
            q = self.filter_form.cleaned_data.get('q')
            if q:
                query = Q(name__icontains=q) | \
                        Q(email__icontains=q) | \
                        Q(phone__icontains=q) | \
                        Q(address__icontains=q)
                
                try:
                    balance_val = Decimal(q.replace(',', ''))
                    query |= Q(balance=balance_val)
                except (ValueError, TypeError, InvalidOperation):
                    pass
                
                qs = qs.filter(query)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = self.filter_form
        context['status'] = self.request.GET.get('status', '')
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_customers(export_format)
        return super().get(request, *args, **kwargs)

    def export_customers(self, format_type):
        customers = self.get_queryset()
        response = generate_customer_list_export(customers, format_type, self.request)
        if response:
            return response
        messages.error(self.request, "Error generating export.")
        return redirect('inventory:customer_list')

@login_required
@require_POST
@permission_required('inventory.add_customerpayment', raise_exception=True)
def customer_payment(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    form = CustomerPaymentForm(request.POST, customer=customer)
    if form.is_valid():
        payment = form.save(commit=False)

        sale_paid = form.cleaned_data.get('sale_paid')
        amount = form.cleaned_data.get('amount')

        if sale_paid:
            # Check for overpayment on a specific invoice
            paid_so_far = sale_paid.payments_received.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            outstanding = sale_paid.total_amount - paid_so_far
            
            if amount > outstanding + Decimal('0.001'): # Add tolerance
                messages.error(request, f"Payment of {amount:,.2f} exceeds the outstanding amount of {outstanding:,.2f} for invoice {sale_paid.receipt_id}.")
                return redirect('inventory:customer_detail', pk=pk)
        else:
            # General payment: Check against total customer balance
            current_balance = customer.get_balance()
            if amount > current_balance + Decimal('0.001'):
                messages.error(request, f"Payment of {amount:,.2f} exceeds the total outstanding balance of {current_balance:,.2f}.")
                return redirect('inventory:customer_detail', pk=pk)

        payment.customer = customer
        payment.recorded_by = request.user
        payment.save()
        messages.success(request, "Payment recorded successfully.")
    else:
        error_str = " ".join([f"{field.replace('_', ' ').title()}: {error}" for field, err_list in form.errors.items() for error in err_list])
        messages.error(request, f"Error recording payment. {error_str if error_str else 'Please check your input.'}")
    return redirect('inventory:customer_detail', pk=pk)

@login_required
@permission_required('inventory.add_customerpayment', raise_exception=True)
def import_ledger_entries(request, pk):
    """Import Ledger Entries (Charges/Payments) from CSV/Excel"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == "POST" and request.FILES.get('csv_file'):
        file_obj = request.FILES['csv_file']
        if not (file_obj.name.lower().endswith('.csv') or file_obj.name.lower().endswith('.xlsx')):
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect('inventory:customer_detail', pk=pk)

        try:
            count_charges, count_payments, errors = inventory_importers.import_ledger_entries_from_file(file_obj, customer, request.user)
            
            if errors:
                for error in errors[:5]:
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.warning(request, f"And {len(errors) - 5} more errors...")
            else:
                if count_payments:
                    messages.success(request, f"Successfully imported {count_payments} payment(s).")
                else:
                    messages.info(request, "Import complete. No new payments were added.")
            
        except Exception as e:
            messages.error(request, f"An unexpected error occurred while processing the file: {e}")
            
        return redirect('inventory:customer_detail', pk=pk)

    instructions = {
        "title": "How to Format Your Payment Data",
        "general":[
            "Fill in your payment data in the downloaded Excel template.",
            "Do NOT change the column headers in the 'Data' sheet.",
            "Delete the sample rows before uploading.",
            "Use YYYY-MM-DD format for dates."
        ],
        "columns":[
            {"name": "Date (*)", "desc": "Payment date in YYYY-MM-DD format.", "example": timezone.now().strftime('%Y-%m-%d')},
            {"name": "Reference", "desc": "Receipt #, Check #, or Transaction Ref.", "example": "'PAY-1001'"},
            {"name": "Description", "desc": "Details about the payment.", "example": "'Partial Payment'"},
            {"name": "Payment (*)", "desc": "Amount paid by customer. Must be a positive number.", "example": "500.00"},
        ],
        "notes":[
            "This import tool is for recording payments only. New charges must be created via POS or SOW."
        ]
    }
    return render(request, 'inventory/ledger_import.html', {'customer': customer, 'instructions': instructions})

@login_required
def download_ledger_template(request):
    """Downloads an Excel template for ledger imports with instructions."""
    wb = Workbook()
    
    # --- Create Instructions Sheet ---
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # Styles
    title_font = Font(name='Calibri', bold=True, size=16, color="1F4E78")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    instruction_header_font = Font(name='Calibri', bold=True, size=12)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Title
    ws_instructions['A1'] = "How to Use This Import Template"
    ws_instructions['A1'].font = title_font
    ws_instructions.merge_cells('A1:D1')

    # General Instructions
    ws_instructions.append([]) # Spacer
    ws_instructions['A3'] = "General Rules"
    ws_instructions['A3'].font = header_font
    ws_instructions.append(["1. Fill in your payment data in the 'Data' sheet."])
    ws_instructions.append(["2. Do NOT change the column headers in the 'Data' sheet."])
    ws_instructions.append(["3. Delete the sample rows in the 'Data' sheet before uploading."])
    ws_instructions.append(["4. Use YYYY-MM-DD format for dates."])
    ws_instructions.append([]) # Spacer
    
    # Column Descriptions
    ws_instructions['A9'] = "Column Guide"
    ws_instructions['A9'].font = header_font
    
    headers =[
        ("Column", "Description", "Example", "Required?"),
        ("Date", "Payment date (YYYY-MM-DD).", timezone.now().strftime('%Y-%m-%d'), "Yes"),
        ("Reference", "Invoice ID (to link) or Check #.", "'JOB-2A505259'", "No"),
        ("Description", "Details about the payment.", "'Partial Payment'", "No"),
        ("Payment", "Amount paid by customer.", "500.00", "Yes"),
    ]
    
    for row_data in headers:
        ws_instructions.append(row_data)

    # Styling for instructions table
    for cell in ws_instructions['A10:D10'][0]:
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row in ws_instructions['A10:D14']:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col_idx, width in enumerate([20, 40, 25, 20], 1):
        ws_instructions.column_dimensions[get_column_letter(col_idx)].width = width
    
    # --- Create Data Sheet ---
    ws_data = wb.create_sheet(title="Data")
    data_headers = ['Date', 'Reference', 'Description', 'Payment']
    ws_data.append(data_headers)
    
    # Sample Data
    ws_data.append([timezone.now().strftime('%Y-%m-%d'), 'PAY-001', 'Partial Payment', 500.00])
    
    # Style header and set column widths
    for i, cell in enumerate(ws_data['1:1'], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws_data.column_dimensions[get_column_letter(i)].width = 25

    ws_data.freeze_panes = 'A2'
    for row in ws_data['A1:D2']:
        for cell in row:
            cell.border = thin_border

    # --- Prepare Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ledger_import_template.xlsx"'
    wb.save(response)
    return response

class CustomerCreateView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'inventory/customer_form.html'
    success_message = "Customer profile for '%(name)s' created successfully."

    def test_func(self):
        return self.request.user.is_superuser
    
    def get_success_url(self):
        return reverse_lazy('inventory:customer_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Add New Customer"
        return context

class CustomerUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'inventory/customer_form.html'
    success_message = "Customer profile for '%(name)s' updated successfully."
    permission_required = 'inventory.change_customer'
    
    def get_success_url(self):
        return reverse_lazy('inventory:customer_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f"Edit: {self.object.name}"
        return context

class CustomerDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Customer
    template_name = 'inventory/customer_detail.html'
    permission_required = 'inventory.view_customer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.object
        
        # Ensure legacy customers have a unique ID generated
        if not customer.customer_id:
            customer.save()
        
        ledger_q = self.request.GET.get('ledger_q', '')

        # 1. Fetch Sales (Credit) with payment status
        sales_qs = customer.purchases.all().select_related('cashier').annotate(
            paid_amount=Coalesce(Sum('payments_received__amount'), Decimal('0.00'))
        ).annotate(
            outstanding=F('total_amount') - F('paid_amount')
        )
        
        # 2. Fetch Payments
        payments_qs = customer.payments.select_related('recorded_by', 'sale_paid')

        if ledger_q:
            query_lower = ledger_q.lower()
            q_sales = Q(receipt_id__icontains=ledger_q) | Q(notes__icontains=ledger_q)
            q_payments = Q(reference_number__icontains=ledger_q) | Q(notes__icontains=ledger_q)

            # Amount Search
            try:
                amount_val = Decimal(ledger_q.replace(',', ''))
                q_sales |= Q(total_amount=amount_val)
                q_payments |= Q(amount=amount_val)
            except (ValueError, TypeError, InvalidOperation):
                pass

            # Keyword Search (Type/Description)
            if 'debt' in query_lower:
                q_sales = Q(payment_method='CREDIT', outstanding__gt=0)
                q_payments = Q(pk__in=[])
            elif 'sale' in query_lower:
                q_sales = Q(payment_method='CASH') | Q(payment_method='CREDIT', outstanding__lte=0)
                q_payments = Q(pk__in=[])
            elif 'payment' in query_lower:
                q_sales = Q(pk__in=[])
                q_payments = Q()
            elif any(k in query_lower for k in ['credit', 'purchase']):
                q_sales = Q()
                # q_payments remains based on text search

            sales_qs = sales_qs.filter(q_sales)
            payments_qs = payments_qs.filter(q_payments)

        payments = list(payments_qs.annotate(
            txn_type=Value('PAYMENT', output_field=models.CharField())
        ).values('payment_date', 'reference_number', 'amount', 'txn_type', 'sale_paid__receipt_id', 'recorded_by__username', 'notes'))
        
        # 3. Combine and Normalize
        ledger =[]
        for s in sales_qs:
            status = "PAID"
            txn_type = 'SALE'
            credit_val = Decimal('0')

            if s.payment_method == 'CREDIT':
                if s.outstanding <= Decimal('0.001'):
                    status = "PAID"
                    txn_type = 'SALE' # Paid off debt becomes Sale
                elif s.paid_amount > 0:
                    status = "PARTIALLY PAID"
                    txn_type = 'DEBT'
                else:
                    status = "UNPAID"
                    txn_type = 'DEBT'
            else:
                # Cash/Card Sales are effectively paid immediately and are Sales
                status = "PAID"
                txn_type = 'SALE'
                credit_val = s.total_amount # Offset debit so balance doesn't increase

            ledger.append({
                'date': s.timestamp,
                'ref': s.receipt_id,
                'description': f'{s.get_payment_method_display()} ({status})',
                'debit': s.total_amount,
                'credit': credit_val,
                'type': txn_type,
                'view_url': reverse('inventory:pos_receipt_detail', kwargs={'receipt_id': s.receipt_id}),
                'user': s.cashier.username if s.cashier else 'N/A'
            })
        for p in payments:
            desc = 'Payment Received'
            view_url = None
            if p['sale_paid__receipt_id']:
                desc += f" (for {p['sale_paid__receipt_id']})"
                view_url = reverse('inventory:pos_receipt_detail', kwargs={'receipt_id': p['sale_paid__receipt_id']})
            if p['notes']:
                desc += f" - {p['notes']}"
            ledger.append({
                'date': p['payment_date'],
                'ref': p['reference_number'],
                'description': desc,
                'debit': 0,
                'credit': p['amount'],
                'type': 'PAYMENT',
                'view_url': view_url,
                'user': p.get('recorded_by__username') or 'N/A'
            })
            
        # 4. Sort by date
        ledger.sort(key=lambda x: x['date'])

        # 5. Calculate Running Balance
        balance = 0
        for entry in ledger:
            balance += (entry['debit'] - entry['credit'])
            entry['balance'] = balance

        # 6. Sort by date descending (Latest first)
        ledger.sort(key=lambda x: x['date'], reverse=True)

        # Pagination for Ledger
        ledger_paginator = Paginator(ledger, 20)
        ledger_page = self.request.GET.get('ledger_page')
        context['ledger'] = ledger_paginator.get_page(ledger_page)
        context['ledger_page_range'] = context['ledger'].paginator.get_elided_page_range(context['ledger'].number, on_each_side=1, on_ends=1)
        
        # Add payment form and financial summary to context
        context['payment_form'] = CustomerPaymentForm(customer=customer)
        current_balance = self.object.get_balance()
        context['current_balance'] = current_balance
        
        # SOW Filtering
        sow_q = self.request.GET.get('sow_q', '')
        sows_qs = self.object.sows.select_related('created_by').all()

        if sow_q:
            q_sow = Q(sow_id__icontains=sow_q) | \
                    Q(hose_type__icontains=sow_q) | \
                    Q(application__icontains=sow_q) | \
                    Q(notes__icontains=sow_q) | \
                    Q(fitting_a__icontains=sow_q) | \
                    Q(fitting_b__icontains=sow_q)
            if sow_q.isdigit():
                q_sow |= Q(id=sow_q)
            sows_qs = sows_qs.filter(q_sow)

        # Pagination for SOW
        sows_paginator = Paginator(sows_qs, 10)
        sow_page = self.request.GET.get('sow_page')
        context['sows'] = sows_paginator.get_page(sow_page)
        context['sows_page_range'] = context['sows'].paginator.get_elided_page_range(context['sows'].number, on_each_side=1, on_ends=1)
        
        context['sow_q'] = sow_q
        
        context['ledger_q'] = ledger_q
        
        # URL Params for Pagination Links (Preserve other filters)
        params = self.request.GET.copy()
        if 'ledger_page' in params: del params['ledger_page']
        if 'sow_page' in params: del params['sow_page']
        context['ledger_query_params'] = params.urlencode()
        
        params = self.request.GET.copy()
        if 'sow_page' in params: del params['sow_page']
        if 'ledger_page' in params: del params['ledger_page']
        context['sow_query_params'] = params.urlencode()

        # Determine active tab
        context['active_tab'] = 'ledger'
        if 'sow_page' in self.request.GET:
            context['active_tab'] = 'sow'
        elif 'ledger_page' in self.request.GET:
            context['active_tab'] = 'ledger'
        elif 'sow_q' in self.request.GET:
            context['active_tab'] = 'sow'

        # URL Params for Exports (Clean all pagination)
        query_params = self.request.GET.copy()
        if 'page' in query_params: query_params.pop('page')
        if 'ledger_page' in query_params: query_params.pop('ledger_page')
        if 'sow_page' in query_params: query_params.pop('sow_page')
        context['query_params'] = query_params.urlencode()

        return context

# --- BILLING STATEMENT EXPORT (Word, Excel, PDF, CSV) ---

@login_required
@permission_required('inventory.view_customer', raise_exception=True)
def export_statement(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    format_type = request.GET.get('format', 'pdf')
    
    ledger_q = request.GET.get('ledger_q', '')

    # --- DATA PREPARATION (aligned with CustomerDetailView) ---
    # 1. Fetch Sales (Credit) with payment status
    sales_qs = customer.purchases.all().select_related('cashier').annotate(
        paid_amount=Coalesce(Sum('payments_received__amount'), Decimal('0.00'))
    ).annotate(
        outstanding=F('total_amount') - F('paid_amount')
    )
    
    # 2. Fetch Payments
    payments_qs = customer.payments.select_related('recorded_by', 'sale_paid')

    if ledger_q:
        query_lower = ledger_q.lower()
        q_sales = Q(receipt_id__icontains=ledger_q) | Q(notes__icontains=ledger_q)
        q_payments = Q(reference_number__icontains=ledger_q) | Q(notes__icontains=ledger_q)

        # Amount Search
        try:
            amount_val = Decimal(ledger_q.replace(',', ''))
            q_sales |= Q(total_amount=amount_val)
            q_payments |= Q(amount=amount_val)
        except (ValueError, TypeError, InvalidOperation):
            pass

        # Keyword Search (Type/Description)
        if 'debt' in query_lower:
            q_sales = Q(payment_method='CREDIT', outstanding__gt=0)
            q_payments = Q(pk__in=[])
        elif 'sale' in query_lower:
            q_sales = Q(payment_method='CASH') | Q(payment_method='CREDIT', outstanding__lte=0)
            q_payments = Q(pk__in=[])
        elif 'payment' in query_lower:
            q_sales = Q(pk__in=[])
            q_payments = Q()
        elif any(k in query_lower for k in['credit', 'purchase']):
            q_sales = Q()
            # q_payments remains based on text search

        sales_qs = sales_qs.filter(q_sales)
        payments_qs = payments_qs.filter(q_payments)

    payments = list(payments_qs.annotate(
        txn_type=Value('PAYMENT', output_field=models.CharField())
    ).values('payment_date', 'reference_number', 'amount', 'txn_type', 'sale_paid__receipt_id', 'recorded_by__username', 'notes'))
    
    # 3. Combine and Normalize
    ledger =[]
    for s in sales_qs:
        status = "PAID"
        credit_val = Decimal('0')

        if s.payment_method == 'CREDIT':
            if s.outstanding <= Decimal('0.001'):
                status = "PAID"
            elif s.paid_amount > 0:
                status = "PARTIALLY PAID"
            else:
                status = "UNPAID"
        else:
            # Cash/Card
            status = "PAID"
            credit_val = s.total_amount

        ledger.append({
            'date': s.timestamp,
            'ref': s.receipt_id,
            'description': f'{s.get_payment_method_display()} ({status})',
            'debit': s.total_amount,
            'credit': credit_val,
            'user': s.cashier.username if s.cashier else 'N/A'
        })
    for p in payments:
        desc = 'Payment Received'
        if p['sale_paid__receipt_id']:
            desc += f" (for {p['sale_paid__receipt_id']})"
        if p['notes']:
            desc += f" - {p['notes']}"
        ledger.append({
            'date': p['payment_date'],
            'ref': p['reference_number'],
            'description': desc,
            'debit': Decimal('0'),
            'credit': p['amount'],
            'user': p.get('recorded_by__username') or 'N/A'
        })
        
    # 4. Sort by date
    ledger.sort(key=lambda x: x['date'])

    # 5. Calculate Running Balance
    balance = Decimal('0')
    for entry in ledger:
        balance += (entry['debit'] - entry['credit'])
        entry['balance'] = balance

    response = generate_customer_statement(customer, ledger, balance, format_type, request)
    if response:
        return response
    return HttpResponse("Error Generating Export", status=500)

# --- PRODUCT MANAGEMENT (UI) ---

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def import_products(request):
    if request.method == 'POST':
        messages.info(request, "Product import functionality is under construction.")
        return redirect('inventory:product_list')
    return render(request, 'inventory/form_import.html', {'title': 'Import Products'})

class ProductListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Product
    context_object_name = 'product_list'
    template_name = 'inventory/product_list.html'
    paginate_by = 12
    permission_required = 'inventory.view_product'

    def get_queryset(self):
        queryset = Product.objects.select_related('category').all()
        form = ProductFilterForm(self.request.GET)
        if form.is_valid():
            query = form.cleaned_data.get('q')
            if query:
                queryset = queryset.filter(Q(name__icontains=query) | Q(sku__icontains=query))
            category = form.cleaned_data.get('category')
            if category:
                queryset = queryset.filter(category=category)
            product_status = form.cleaned_data.get('product_status')
            if product_status:
                queryset = queryset.filter(status=product_status)
            
            stock_status = form.cleaned_data.get('stock_status')
            if stock_status:
                if stock_status == 'in_stock':
                    queryset = queryset.filter(quantity__gt=10)
                elif stock_status == 'low_stock':
                    queryset = queryset.filter(quantity__gt=0, quantity__lte=10)
                elif stock_status == 'out_of_stock':
                    queryset = queryset.filter(quantity=0)

            sort_by = form.cleaned_data.get('sort_by')
            if sort_by:
                queryset = queryset.order_by(sort_by)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductFilterForm(self.request.GET)
        context['category_form'] = CategoryCreateForm()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

class ProductDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Product
    template_name = 'inventory/product_detail.html'
    context_object_name = 'product'
    permission_required = 'inventory.view_product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = StockTransaction.objects.filter(product=self.object).order_by('-timestamp')[:10]
        context['transaction_form'] = StockOutForm()
        context['refund_form'] = RefundForm(product=self.object)
        return context
    
    def post(self, request, *args, **kwargs):
        with transaction.atomic():
            product_object = Product.objects.select_for_update().get(pk=self.get_object().pk)
            form = StockOutForm(request.POST)
            
            if form.is_valid():
                transaction_obj = form.save(commit=False)
                transaction_obj.product = product_object
                transaction_obj.user = request.user
                transaction_obj.transaction_type = 'OUT'
                
                quantity = form.cleaned_data.get('quantity')
                if product_object.quantity < quantity:
                    messages.error(request, f'Cannot stock out more than available ({product_object.quantity}).')
                    return redirect(product_object.get_absolute_url())
                
                product_object.quantity -= quantity
                product_object.save()
                
                transaction_obj.selling_price = product_object.price if transaction_obj.transaction_reason == 'SALE' else None
                transaction_obj.save()
                messages.success(request, "Stock Out recorded successfully.")
            else:
                messages.error(request, "Error recording transaction.")
        return redirect(product_object.get_absolute_url())

@login_required
@require_POST
@permission_required('inventory.can_adjust_stock', raise_exception=True)
def product_refund(request, slug):
    product = get_object_or_404(Product, slug=slug)
    form = RefundForm(request.POST, product=product)
    
    if form.is_valid():
        sale = form.cleaned_data['pos_sale']
        receipt_id = sale.receipt_id
        quantity = form.cleaned_data['quantity']
        notes = form.cleaned_data.get('notes')

        # 2. Verify Product was in that Receipt
        sold_items = StockTransaction.objects.filter(
            pos_sale=sale,
            product=product,
            transaction_type='OUT',
            transaction_reason=StockTransaction.TransactionReason.SALE
        )
        total_sold = sold_items.aggregate(total=Sum('quantity'))['total'] or 0

        if total_sold == 0:
            messages.error(request, f"Product '{product.name}' was not found in Receipt {receipt_id}.")
            return redirect(product.get_absolute_url())

        # 3. Check Previous Returns (Prevent over-refunding)
        returned_items = StockTransaction.objects.filter(
            pos_sale=sale,
            product=product,
            transaction_type='IN',
            transaction_reason=StockTransaction.TransactionReason.RETURN
        )
        total_returned = returned_items.aggregate(total=Sum('quantity'))['total'] or 0

        if (total_returned + quantity) > total_sold:
            remaining = total_sold - total_returned
            messages.error(request, f"Cannot refund {quantity}. Only {remaining} items eligible for return from this receipt.")
            return redirect(product.get_absolute_url())

        with transaction.atomic():
            StockTransaction.objects.create(
                product=product,
                transaction_type='IN',
                transaction_reason=StockTransaction.TransactionReason.RETURN,
                quantity=quantity,
                user=request.user,
                selling_price=product.price, 
                pos_sale=sale,
                notes=f"Refund for Receipt {receipt_id}: {notes}"
            )
            product.quantity += quantity
            product.save()
            messages.success(request, f"Refund processed. {quantity} items returned from Receipt {receipt_id}.")
    else:
        for field, errors in form.errors.items():
            messages.error(request, f"{field}: {', '.join(errors)}")
            
    return redirect(product.get_absolute_url())

class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Product
    form_class = ProductCreateForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('inventory:product_list')
    success_message = "Product was created successfully!"
    permission_required = 'inventory.add_product'

class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Product
    form_class = ProductUpdateForm
    template_name = 'inventory/product_form.html'
    success_message = "Product was updated successfully!"
    permission_required = 'inventory.change_product'
    
    def get_success_url(self):
        return self.object.get_absolute_url()

class ProductDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('inventory:product_list')
    permission_required = 'inventory.delete_product'

# --- POINT OF SALE (POS) SYSTEM ---

def get_walkin_customer():
    """Helper to get or create the default Walk-in Customer."""
    customer, created = Customer.objects.get_or_create(
        name="Walk-in Customer",
        defaults={
            'email': '',
            'phone': '',
            'address': 'Store Counter',
            'tax_id': '',
            'credit_limit': 0
        }
    )
    return customer

@login_required
@permission_required('inventory.add_possale', raise_exception=True)
def pos_dashboard(request):
    # Products
    active_products_qs = Product.objects.filter(status=Product.Status.ACTIVE, quantity__gt=0).values(
        'id', 'name', 'sku', 'price', 'quantity', 'category__name', 'image'
    )

    # Manually process to add the full image URL
    products_list =[]
    for p in active_products_qs:
        image_url = None
        if p.get('image'):
            # Construct the full URL path for the template
            image_url = f"{settings.MEDIA_URL}{p['image']}"
        p['image_url'] = image_url
        products_list.append(p)

    products_json = json.dumps(products_list, cls=DjangoJSONEncoder)
    
    # Customers
    customers = Customer.objects.values('id', 'name')
    customers_json = json.dumps(list(customers), cls=DjangoJSONEncoder)
    
    # Get pre-selected customer from URL
    preselected_customer_id = request.GET.get('customer_id')

    # Ensure Walk-in Customer exists
    walkin_customer = get_walkin_customer()

    context = {
        'page_title': 'Point of Sale',
        'products_json': products_json,
        'customers_json': customers_json,
        'preselected_customer_id': preselected_customer_id,
        'walkin_customer': walkin_customer,
    }
    return render(request, 'inventory/pos.html', context)

@login_required
@permission_required('inventory.add_hydraulicsow', raise_exception=True)
def pos_sow_create(request):
    walkin = get_walkin_customer()
    # Redirect to SOW create with next=pos_dashboard
    url = reverse('inventory:hydraulic_sow_create', kwargs={'pk': walkin.pk})
    next_url = reverse('inventory:pos_dashboard')
    return redirect(f"{url}?next={next_url}")

@login_required
@require_POST
@permission_required('inventory.add_possale', raise_exception=True)
def pos_checkout(request):
    try:
        data = json.loads(request.body)
        items = data.get('items',[])
        
        # New: Customer and Payment Method Logic
        customer_id = data.get('customer_id') 
        payment_method = data.get('payment_method', 'CASH') # CASH, CREDIT, CARD
        
        raw_amount = data.get('amount_paid')
        amount_paid = Decimal(str(raw_amount)) if raw_amount else Decimal('0')

        if not items:
            return JsonResponse({'status': 'error', 'message': 'Cart is empty'}, status=400)

        customer = None
        if customer_id:
            try:
                customer = Customer.objects.get(pk=customer_id)
            except Customer.DoesNotExist:
                pass

        # Default to Walk-in Customer if no specific customer is selected
        if not customer:
            customer = get_walkin_customer()

        # Calculate Total Cost first to validate Credit Limit
        total_calculated_cost = Decimal('0')
        item_objects =[]
        
        # Pre-validation Loop
        for item in items:
            product = Product.objects.get(pk=item.get('id'))
            qty = int(item.get('qty'))

            # Allow for custom price override (requested lower price)
            custom_price = item.get('price')
            original_price = item.get('original_price')
            override_reason = item.get('override_reason')

            if custom_price is not None:
                try:
                    sell_price = Decimal(str(custom_price))
                except (ValueError, TypeError):
                    sell_price = product.price
            else:
                sell_price = product.price

            if original_price is not None:
                original_price_decimal = Decimal(str(original_price))
            else:
                original_price_decimal = product.price # Fallback

            if product.quantity < qty:
                raise ValueError(f"Insufficient stock for {product.name}")
            
            total_calculated_cost += (sell_price * qty)
            item_objects.append({
                'product': product,
                'qty': qty,
                'price': sell_price,
                'original_price': original_price_decimal,
                'override_reason': override_reason
            })

        # Credit Validation
        if payment_method == 'CREDIT':
            if not customer:
                 return JsonResponse({'status': 'error', 'message': 'Customer required for credit sales'}, status=400)
            
            # If credit, immediate payment is 0
            amount_paid = Decimal('0') 
        else: # CASH, CARD, GCASH, BANK
             if amount_paid < total_calculated_cost:
                 return JsonResponse({'status': 'error', 'message': 'Payment amount is less than total cost.'}, status=400)

        receipt_id = f"REC-{uuid.uuid4().hex[:8].upper()}"
        
        # Check if any item has a price override
        sale_has_override = any(
            item.get('price') is not None and item.get('original_price') is not None and
            Decimal(str(item.get('price'))) < Decimal(str(item.get('original_price')))
            for item in items
        )
        
        with transaction.atomic():
            # 1. Create Sale Header
            sale_record = POSSale.objects.create(
                receipt_id=receipt_id,
                cashier=request.user,
                customer=customer,
                payment_method=payment_method,
                total_amount=total_calculated_cost, 
                amount_paid=amount_paid,
                change_given=(amount_paid - total_calculated_cost) if payment_method != 'CREDIT' else 0,
                has_price_override=sale_has_override
            )

            receipt_items_response =[]
            
            # 2. Process Items
            for item_obj in item_objects:
                product = item_obj['product']
                sell_qty = item_obj['qty']
                sell_price = item_obj['price']
                original_price = item_obj['original_price']
                override_reason = item_obj.get('override_reason')
                
                # Lock row
                product = Product.objects.select_for_update().get(pk=product.id)

                # Log price override if price was lowered
                if sell_price < original_price:
                    PriceOverrideLog.objects.create(
                        pos_sale=sale_record,
                        product=product,
                        salesman=request.user,
                        original_price=original_price,
                        override_price=sell_price,
                        reason=override_reason
                    )

                product.quantity -= sell_qty
                product.save()
                
                line_total = sell_qty * sell_price
                
                txn_notes = f"POS Sale: {receipt_id} ({payment_method})"
                if sell_price < original_price:
                    txn_notes += f" | Price Override: {original_price:,.2f} -> {sell_price:,.2f}"
                    if override_reason:
                        txn_notes += f"[Reason: {override_reason}]"

                StockTransaction.objects.create(
                    product=product,
                    transaction_type='OUT',
                    transaction_reason=StockTransaction.TransactionReason.SALE,
                    quantity=sell_qty,
                    selling_price=sell_price,
                    user=request.user,
                    pos_sale=sale_record,
                    notes=txn_notes
                )
                
                receipt_items_response.append({
                    'name': product.name,
                    'qty': sell_qty,
                    'price': f"{sell_price:,.2f}",
                    'total': f"{line_total:,.2f}"
                })

            return JsonResponse({
                'status': 'success', 
                'receipt_id': receipt_id,
                'date': sale_record.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                'customer_name': customer.name if customer else 'Walk-in',
                'items': receipt_items_response,
                'total': f"{total_calculated_cost:,.2f}",
                'amount_paid': f"{amount_paid:,.2f}",
                'change': f"{sale_record.change_given:,.2f}"
            })

    except Product.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

class POSHistoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = POSSale
    template_name = 'inventory/pos_history.html'
    context_object_name = 'sales'
    paginate_by = 20
    permission_required = 'inventory.view_possale'
    
    def get_queryset(self):
        qs = POSSale.objects.filter(
            Q(customer__name="Walk-in Customer") | Q(customer__isnull=True)
        ).select_related('cashier', 'customer').order_by('-timestamp')
        
        txn_type = self.request.GET.get('type')
        if txn_type == 'REC':
            qs = qs.filter(receipt_id__startswith='REC')
        elif txn_type == 'JOB':
            qs = qs.filter(Q(receipt_id__startswith='JOB') | Q(receipt_id__startswith='SOW'))
            
        q = self.request.GET.get('q')
        if q:
            query = Q(receipt_id__icontains=q) | \
                    Q(customer__name__icontains=q) | \
                    Q(cashier__username__icontains=q)
            try:
                amount_val = Decimal(q.replace(',', ''))
                query |= Q(total_amount=amount_val)
            except (ValueError, TypeError, InvalidOperation):
                pass
            qs = qs.filter(query)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        context['type'] = self.request.GET.get('type', '')
        
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        
        return context

class POSReceiptDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = POSSale
    context_object_name = 'sale'
    permission_required = 'inventory.view_possale'
    template_name = 'inventory/pos_receipt.html'
    
    def get_object(self):
        return get_object_or_404(POSSale, receipt_id=self.kwargs['receipt_id'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Fetch product items
        items_qs = self.object.items.select_related('product').annotate(
            line_total=ExpressionWrapper(F('quantity') * F('selling_price'), output_field=DecimalField())
        )
        items = list(items_qs)

        # Fetch associated Hydraulic SOW specifications if this is a job receipt
        is_job = self.object.receipt_id.startswith(('JOB-', 'SOW-'))
        if is_job:
            sow = HydraulicSow.objects.filter(sow_id=self.object.receipt_id).select_related('customer', 'created_by').first()
            if sow:
                context['sow'] = sow
                
                # Synthesize a service description item if no products are linked
                if not items:
                    service_desc = f"Hydraulic Service: {sow.hose_type} | Ø {sow.diameter} | {sow.fitting_a}/{sow.fitting_b}"
                    if sow.application: service_desc += f" ({sow.application})"
                    
                    items.append({
                        'product': {'name': service_desc, 'sku': sow.sow_id},
                        'quantity': 1,
                        'selling_price': self.object.total_amount,
                        'line_total': self.object.total_amount
                    })

        context['items'] = items
        return context

def get_product_price(request, product_id):
    """API endpoint for admin JS to fetch product price."""
    from django.http import JsonResponse
    try:
        product = Product.objects.get(pk=product_id)
        return JsonResponse({'price': str(product.price)})
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

# --- ANALYTICS & REPORTS ---

@method_decorator(xframe_options_exempt, name='dispatch')
class ReportingView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'inventory/reporting.html'
    permission_required = 'inventory.can_view_reports'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Report Generation"
        context['transaction_report_form'] = TransactionReportForm(self.request.GET or None)
        return context

    def get(self, request, *args, **kwargs):
        export_type = request.GET.get('export')
        if export_type == 'inventory_pdf':
            return self.export_inventory_pdf(request)
        elif export_type == 'inventory_excel':
            return self.export_inventory_excel(request)
        elif export_type == 'transaction_pdf':
            return self.export_transactions_pdf(request)
        return super().get(request, *args, **kwargs)

    def export_inventory_excel(self, request):
        from .exports import generate_inventory_snapshot_export
        products = Product.objects.select_related('category').all().annotate(
            total_value=ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())
        ).order_by('name')

        total_inventory_value = products.aggregate(total=Sum('total_value'))['total'] or Decimal('0.00')
        total_items = products.aggregate(total=Sum('quantity'))['total'] or 0

        return generate_inventory_snapshot_export(products, total_inventory_value, total_items, request)

    def export_inventory_pdf(self, request):
        products = Product.objects.select_related('category').all().annotate(
            total_value=ExpressionWrapper(F('quantity') * F('price'), output_field=DecimalField())
        ).order_by('name')
        
        total_inventory_value = products.aggregate(total=Sum('total_value'))['total'] or Decimal('0.00')
        total_items = products.aggregate(total=Sum('quantity'))['total'] or 0

        context = {
            'products': products,
            'total_inventory_value': total_inventory_value,
            'total_items': total_items,
            'today': timezone.now(),
        }

        pdf = render_to_pdf('inventory/stock_snapshot_pdf.html', context, request=request)
        
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Stock_Snapshot_{timezone.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        messages.error(request, "Could not generate PDF report. Please try again.")
        return redirect('inventory:reporting')

    def export_transactions_pdf(self, request):
        form = TransactionReportForm(request.GET)
        
        start_date, end_date = None, None
        transactions = StockTransaction.objects.select_related('product').all()

        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            if start_date:
                start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
                transactions = transactions.filter(timestamp__gte=start_dt)
            if end_date:
                end_dt = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
                transactions = transactions.filter(timestamp__lte=end_dt)
        
        # 1. Consolidate main aggregations and count into ONE query
        # Use direct F expressions instead of annotations to avoid subqueries in some DBs
        summary = transactions.aggregate(
            gross_sales=Sum(F('quantity') * F('selling_price'), filter=Q(transaction_reason=StockTransaction.TransactionReason.SALE)),
            total_refunds=Sum(F('quantity') * F('selling_price'), filter=Q(transaction_reason=StockTransaction.TransactionReason.RETURN)),
            total_items_sold=Sum('quantity', filter=Q(transaction_reason=StockTransaction.TransactionReason.SALE)),
            total_count=Count('id')
        )
        
        gross_sales = summary['gross_sales'] or Decimal('0.00')
        total_refunds = summary['total_refunds'] or Decimal('0.00')
        net_revenue = gross_sales - total_refunds
        total_items_sold = summary['total_items_sold'] or 0
        total_count = summary['total_count'] or 0

        # 2. Summaries (Reason-based)
        inflow_summary = transactions.filter(transaction_type='IN').values('transaction_reason').annotate(total_qty=Sum('quantity')).order_by('-total_qty')
        
        loss_summary = transactions.filter(
            transaction_reason__in=[StockTransaction.TransactionReason.DAMAGE, StockTransaction.TransactionReason.INTERNAL]
        ).values('transaction_reason').annotate(
            total_qty=Sum('quantity'), 
            total_val=Sum(F('quantity') * F('product__price'))
        ).order_by('-total_val')

        top_sellers = transactions.filter(
            transaction_reason=StockTransaction.TransactionReason.SALE
        ).values('product__name').annotate(
            total_quantity_sold=Sum('quantity')
        ).order_by('-total_quantity_sold')[:5]

        context = {
            'start_date': start_date, 'end_date': end_date,
            'gross_sales': gross_sales, 'total_refunds': total_refunds, 'net_revenue': net_revenue,
            'total_items_sold': total_items_sold, 'inflow_summary': inflow_summary,
            'loss_summary': loss_summary, 'top_sellers': top_sellers, 'today': timezone.now(),
            'total_count': total_count
        }

        pdf = render_to_pdf('inventory/transaction_report_pdf.html', context, request=request)
        
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Stock_Movement_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
            if 'preview' in request.GET:
                response['Content-Disposition'] = f'inline; filename="{filename}"'
            else:
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        messages.error(request, "Could not generate PDF report. Please try again.")
        return redirect('inventory:reporting')

@login_required
@permission_required('inventory.can_view_reports', raise_exception=True)
def analytics_dashboard(request):
    # 1. Date Filtering
    today = timezone.now().date()
    
    # Default to current month/year
    default_month = str(today.month)
    default_year = str(today.year)
    
    data = request.GET.copy()
    if not request.GET:
        data['month'] = default_month
        data['year'] = default_year
    
    filter_form = AnalyticsFilterForm(data)
    
    # Determine Date Range & Period Name
    start_date = today.replace(day=1)
    end_date = today
    period_name = start_date.strftime('%B %Y')

    if filter_form.is_valid():
        m = filter_form.cleaned_data.get('month')
        y = filter_form.cleaned_data.get('year')
        
        if y:
            year_val = int(y)
            if m:
                month_val = int(m)
                start_date = datetime(year_val, month_val, 1).date()
                # Calculate last day of month
                if month_val == 12:
                    end_date = datetime(year_val + 1, 1, 1).date() - timedelta(days=1)
                else:
                    end_date = datetime(year_val, month_val + 1, 1).date() - timedelta(days=1)
                period_name = start_date.strftime('%B %Y')
            else:
                # Full Year
                start_date = datetime(year_val, 1, 1).date()
                end_date = datetime(year_val, 12, 31).date()
                period_name = f"Year {y}"

    # Make end_date inclusive (end of the day)
    start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    # 2. Base Querysets
    pos_sales = POSSale.objects.filter(timestamp__range=[start_dt, end_dt])
    stock_txns = StockTransaction.objects.filter(timestamp__range=[start_dt, end_dt])
    expenses_qs = Expense.objects.filter(expense_date__range=[start_date, end_date])
    
    # 3. KPI Calculations
    # Optimized: Use conditional aggregation to reduce DB queries
    stock_metrics = stock_txns.aggregate(
        refunds_val=Sum(
            F('quantity') * F('selling_price'),
            filter=Q(transaction_type='IN', transaction_reason=StockTransaction.TransactionReason.RETURN)
        ),
        units_sold=Sum(
            'quantity',
            filter=Q(transaction_type='OUT', transaction_reason=StockTransaction.TransactionReason.SALE)
        ),
        loss_val=Sum(
            F('quantity') * F('product__price'),
            filter=Q(transaction_reason=StockTransaction.TransactionReason.DAMAGE)
        ),
        refunds_count=Count(
            'id',
            filter=Q(transaction_type='IN', transaction_reason=StockTransaction.TransactionReason.RETURN)
        ),
        damages_count=Count(
            'id',
            filter=Q(transaction_reason=StockTransaction.TransactionReason.DAMAGE)
        )
    )
    
    gross_sales_val = pos_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_expenses = expenses_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_refunds_val = stock_metrics['refunds_val'] or Decimal('0.00')
    total_units = stock_metrics['units_sold'] or 0
    total_loss = stock_metrics['loss_val'] or Decimal('0.00')
    
    total_refunds_count = stock_metrics['refunds_count'] or 0
    total_damages_count = stock_metrics['damages_count'] or 0
    
    charges_qs = pos_sales.filter(payment_method='CREDIT')
    charges_count = charges_qs.count()
    total_charges_val = charges_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')

    # Correct Financial Logic
    # Gross Sales = Total money from sales (before refunds)
    # Net Revenue = Gross Sales - Refunds
    # Net Income  = Net Revenue - Expenses
    
    net_revenue_val = gross_sales_val - total_refunds_val
    net_income = net_revenue_val - total_expenses
    

    # 4. Chart Data Preparation
    # A. Sales by Category
    # Optimized: Group by ID to avoid joins during aggregation
    cat_qs = stock_txns.filter(
        transaction_type='OUT', transaction_reason=StockTransaction.TransactionReason.SALE
    ).values('product__category').annotate(sales=Sum(F('quantity') * F('selling_price'))).order_by('-sales')
    
    cat_ids = [item['product__category'] for item in cat_qs if item['product__category']]
    categories = Category.objects.filter(id__in=cat_ids).in_bulk()
    
    cat_labels =[]
    cat_values = []
    for item in cat_qs:
        cat_id = item['product__category']
        name = categories[cat_id].name if cat_id in categories else 'Uncategorized'
        cat_labels.append(name)
        cat_values.append(float(item['sales'] or 0))
    
    # B. Top 5 Best Selling Products
    # Optimized: Group by ID to avoid joins during aggregation
    prod_qs = stock_txns.filter(
        transaction_type='OUT', transaction_reason=StockTransaction.TransactionReason.SALE
    ).values('product').annotate(sales=Sum(F('quantity') * F('selling_price'))).order_by('-sales')[:5]
    
    prod_ids = [item['product'] for item in prod_qs]
    products = Product.objects.filter(id__in=prod_ids).in_bulk()
    
    prod_labels =[]
    prod_values = []
    for item in prod_qs:
        prod_id = item['product']
        if prod_id in products:
            prod_labels.append(products[prod_id].name)
            prod_values.append(float(item['sales'] or 0))
    
    # C. Expenses by Category
    exp_cat_qs = expenses_qs.values('category__name').annotate(total=Sum('amount')).order_by('-total')
    exp_cat_labels =[item['category__name'] or 'Uncategorized' for item in exp_cat_qs]
    exp_cat_values = [float(item['total'] or 0) for item in exp_cat_qs]
    
    # D. Financial Trend (Sales vs Expenses)
    sales_trend = pos_sales.annotate(date=TruncDate('timestamp')).values('date').annotate(daily_total=Sum('total_amount')).order_by('date')
    sales_map = {item['date']: item['daily_total'] for item in sales_trend}

    exp_trend = expenses_qs.values('expense_date').annotate(daily_total=Sum('amount')).order_by('expense_date')
    exp_map = {item['expense_date']: item['daily_total'] for item in exp_trend}

    all_dates = sorted(list(set(list(sales_map.keys()) + list(exp_map.keys()))))
    
    trend_labels = [d.strftime('%b %d') for d in all_dates]
    trend_sales_values = [float(sales_map.get(d, 0)) for d in all_dates]
    trend_expense_values =[float(exp_map.get(d, 0)) for d in all_dates]
    
    # E. Sales Breakdown (Cash, Charges, Hydraulic)
    
    # 1. Cash (Cash + Card)
    cash_sales = pos_sales.filter(
        payment_method__in=['CASH', 'CARD', 'GCASH', 'BANK']
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # 2. Hydraulic Jobs (Credit) - Identified by 'JOB-' or 'SOW-' prefix
    hydraulic_sales = pos_sales.filter(
        Q(receipt_id__startswith='JOB') | Q(receipt_id__startswith='SOW'),
        payment_method='CREDIT'
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # 3. Charges (Credit excluding Hydraulic)
    other_charges = pos_sales.filter(
        payment_method='CREDIT'
    ).exclude(
        Q(receipt_id__startswith='JOB') | Q(receipt_id__startswith='SOW')
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    pay_labels = ['Cash', 'Charges', 'Hydraulic Jobs']
    pay_values =[float(cash_sales), float(other_charges), float(hydraulic_sales)]

    context = {
        'filter_form': filter_form,
        'start_date': start_date,
        'end_date': end_date,
        'period_name': period_name,
        
        # KPIs
        'total_revenue': net_revenue_val, # Template label is "Net Revenue"
        'gross_sales': gross_sales_val,   # Template label is "Gross"
        'total_expenses': total_expenses,
        'net_income': net_income,
        'total_units': total_units,
        'total_refunds': total_refunds_val,
        'total_loss': total_loss,
        'charges_count': charges_count,
        'total_charges_val': total_charges_val,
        'refunds_count': total_refunds_count,
        'damages_count': total_damages_count,
        
        # Charts (JSON)
        'cat_labels': cat_labels,
        'cat_values': cat_values,
        'exp_cat_labels': exp_cat_labels,
        'exp_cat_values': exp_cat_values,
        'prod_labels': prod_labels,
        'prod_values': prod_values,
        'trend_labels': trend_labels,
        'trend_sales_values': trend_sales_values,
        'trend_expense_values': trend_expense_values,
        'pay_labels': pay_labels,
        'pay_values': pay_values,
    }
    return render(request, 'inventory/analytics.html', context)

# --- PURCHASE ORDERS & SUPPLIERS (Existing) ---

class PurchaseOrderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = PurchaseOrder
    template_name = 'inventory/purchaseorder_list.html'
    context_object_name = 'po_list'
    paginate_by = 20
    permission_required = 'inventory.view_purchaseorder'
    
    def get_queryset(self):
        queryset = PurchaseOrder.objects.select_related('supplier').order_by('-order_date')
        self.filter_form = PurchaseOrderFilterForm(self.request.GET)
        if self.filter_form.is_valid():
            if self.filter_form.cleaned_data.get('supplier'):
                queryset = queryset.filter(supplier=self.filter_form.cleaned_data['supplier'])
            if self.filter_form.cleaned_data.get('status'):
                queryset = queryset.filter(status=self.filter_form.cleaned_data['status'])
            if self.filter_form.cleaned_data.get('start_date'):
                queryset = queryset.filter(order_date__date__gte=self.filter_form.cleaned_data['start_date'])
            if self.filter_form.cleaned_data.get('end_date'):
                queryset = queryset.filter(order_date__date__lte=self.filter_form.cleaned_data['end_date'])
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = self.filter_form
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

class PurchaseOrderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = PurchaseOrder
    template_name = 'inventory/purchaseorder_detail.html'
    context_object_name = 'po'
    permission_required = 'inventory.view_purchaseorder'

@login_required
@permission_required('inventory.change_purchaseorder', raise_exception=True)
def receive_purchase_order(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST' and po.status == 'COMPLETED':
        po.complete_order(request.user)
        messages.success(request, f"Stock from {po.order_id} added.")
    return redirect('inventory:purchaseorder_list')

class SupplierListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Supplier
    template_name = 'inventory/supplier_list.html'
    paginate_by = 10
    permission_required = 'inventory.view_supplier'

    def get_queryset(self):
        queryset = super().get_queryset().order_by('name')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(name__icontains=q) |
                Q(contact_person__icontains=q) |
                Q(email__icontains=q) |
                Q(supplier_id__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.request.GET.get('q', '')
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

class SupplierDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Supplier
    template_name = 'inventory/supplier_detail.html'
    context_object_name = 'supplier'
    permission_required = 'inventory.view_supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.object
        
        po_list = supplier.purchase_orders.all().order_by('-order_date')
        
        paginator = Paginator(po_list, 15)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context['purchase_orders'] = page_obj
        context['is_paginated'] = page_obj.has_other_pages()
        context['page_obj'] = page_obj
        context['po_page_range'] = paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1)
        
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        export_format = request.GET.get('export')
        if export_format:
            return self.export_deliveries(export_format)
        
        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_purchase_orders(self):
        return self.object.purchase_orders.prefetch_related('items', 'items__product').order_by('-order_date')

    def export_deliveries(self, format_type):
        supplier = self.object
        purchase_orders = self.get_purchase_orders()
        response = generate_supplier_deliveries_export(supplier, purchase_orders, format_type, self.request)
        if response:
            return response
        messages.error(self.request, "Error generating export.")
        return redirect('inventory:supplier_detail', pk=supplier.pk)

@login_required
@permission_required('inventory.add_purchaseorder', raise_exception=True)
def import_supplier_deliveries(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST" and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        if not (csv_file.name.endswith('.csv') or csv_file.name.endswith('.xlsx')):
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect('inventory:supplier_detail', pk=pk)
        try:
            data =[]
            if csv_file.name.endswith('.csv'):
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                reader.fieldnames =[name.strip().lower().replace(' ', '_') for name in reader.fieldnames]
                data = list(reader)
            elif csv_file.name.endswith('.xlsx'):
                wb = load_workbook(csv_file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers =[str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
                    for row in rows[1:]:
                        row_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
                        data.append(row_dict)

            created_pos = {}
            items_added = 0
            with transaction.atomic():
                for row in data:
                    po_id = row.get('po_id')
                    product_sku = row.get('product_sku')
                    
                    # Handle Excel int/float vs CSV string
                    qty_val = row.get('quantity', 0)
                    if qty_val is None: qty_val = 0
                    quantity = int(qty_val)
                    
                    price_val = row.get('price', 0)
                    if price_val is None: price_val = 0
                    price = Decimal(str(price_val))

                    if not all([po_id, product_sku, quantity > 0]):
                        continue
                    if po_id not in created_pos:
                        po, created = PurchaseOrder.objects.get_or_create(order_id=po_id, defaults={'supplier': supplier, 'status': 'PENDING'})
                        if not created and po.supplier != supplier:
                            raise ValueError(f"Purchase Order ID {po_id} already exists for another supplier.")
                        created_pos[po_id] = po
                    po = created_pos[po_id]
                    try:
                        product = Product.objects.get(sku=product_sku)
                    except Product.DoesNotExist:
                        messages.warning(request, f"Product with SKU '{product_sku}' not found. Skipping item in PO {po_id}.")
                        continue
                    PurchaseOrderItem.objects.create(purchase_order=po, product=product, quantity=quantity, price=price)
                    items_added += 1
            messages.success(request, f"Successfully imported {items_added} items across {len(created_pos)} Purchase Orders.")
        except ValueError as e:
            messages.error(request, f"Data error: {e}")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('inventory:supplier_detail', pk=pk)
    return render(request, 'inventory/supplier_deliveries_import.html', {'supplier': supplier})

# --- API VIEWS (DRF) ---

@extend_schema(tags=['Inventory Management'])
class CategoryViewSet(viewsets.ModelViewSet):
    http_method_names =['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class ProductViewSet(viewsets.ModelViewSet):
    http_method_names = ['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'sku']

# --- NEW VIEWSETS ---

@extend_schema(tags=['Customers & Billing'])
class CustomerViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing Customers.
    Provides full CRUD functionality for customer profiles.
    The 'balance' is a read-only calculated field.
    """
    http_method_names =['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = Customer.objects.all().order_by('name')
    serializer_class = CustomerSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends =[filters.SearchFilter]
    search_fields = ['name', 'customer_id', 'email', 'phone']

@extend_schema(tags=['Customers & Billing'])
class CustomerPaymentViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing Customer Payments.
    Allows creating, viewing, and managing payments.
    'recorded_by' is automatically set to the logged-in user on creation.
    """
    http_method_names = ['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = CustomerPayment.objects.select_related('customer', 'recorded_by', 'sale_paid').all()
    serializer_class = CustomerPaymentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields =['customer__name', 'reference_number', 'sale_paid__receipt_id']

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)

@extend_schema(tags=['Customers & Billing'])
class HydraulicSowViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing Hydraulic Scope of Work (SOW) jobs.
    'created_by' is automatically set to the logged-in user on creation.
    """
    http_method_names =['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = HydraulicSow.objects.select_related('customer', 'created_by').all()
    serializer_class = HydraulicSowSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['sow_id', 'customer__name', 'application', 'hose_type']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

@extend_schema(tags=['Point of Sale'])
class POSSaleViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing Point of Sale (POS) transactions.
    This is a read-only endpoint as sales are created through the POS checkout process.
    """
    queryset = POSSale.objects.select_related('cashier', 'customer').prefetch_related('items', 'items__product').all()
    serializer_class = POSSaleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields =['receipt_id', 'customer__name', 'cashier__username']

@extend_schema(tags=['Expenses'])
class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    """API endpoint for managing Expense Categories."""
    http_method_names =['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes =[permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

@extend_schema(tags=['Expenses'])
class ExpenseViewSet(viewsets.ModelViewSet):
    """API endpoint for managing Expenses. 'recorded_by' is automatically set to the logged-in user on creation."""
    http_method_names = ['get', 'post', 'put', 'delete', 'head', 'options']
    queryset = Expense.objects.select_related('category', 'recorded_by').all()
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends =[filters.SearchFilter]
    search_fields = ['description', 'category__name']

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)

# --- AJAX HELPERS ---

@login_required
@require_POST
def add_category_ajax(request):
    form = CategoryCreateForm(request.POST)
    if form.is_valid():
        cat = form.save()
        return JsonResponse({'status': 'success', 'category': {'id': cat.id, 'name': cat.name}})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
@require_POST
def add_expense_category_ajax(request):
    name = request.POST.get('name')
    if name:
        cat, created = ExpenseCategory.objects.get_or_create(name=name)
        return JsonResponse({'status': 'success', 'category': {'id': cat.id, 'name': cat.name}})
    return JsonResponse({'status': 'error', 'message': 'Category name is required.'}, status=400)

@login_required
def search_products(request):
    """AJAX endpoint for searching products by name or SKU."""
    query = request.GET.get('q', '')
    if query:
        products = Product.objects.filter(
            Q(name__icontains=query) | Q(sku__icontains=query)
        ).filter(status=Product.Status.ACTIVE).values('id', 'name', 'sku', 'price', 'quantity')[:20]
        return JsonResponse({'results': list(products)})
    return JsonResponse({'results':[]})

@login_required
def sales_chart_data(request):
    # Removed Hour and Minute sales as requested, defaulting to daily view
    now = timezone.now()
    start_time = now - timedelta(days=30)
    trunc_func = TruncDate('timestamp')
    date_format = '%b %d'

    # Fetch POS Sales grouped by Date and Payment Method
    sales_qs = POSSale.objects.filter(timestamp__gte=start_time).annotate(
        period_group=trunc_func
    ).values('period_group', 'payment_method').annotate(
        total=Sum('total_amount')
    ).order_by('period_group')
    
    # Organize data into dictionaries
    sales_by_date = {}
    charges_by_date = {}

    for entry in sales_qs:
        d = entry['period_group']
        if isinstance(d, datetime): d = d.date()
        
        amount = float(entry['total'])
        if entry['payment_method'] == 'CREDIT':
            charges_by_date[d] = charges_by_date.get(d, 0) + amount
        else:
            # Group CASH and CARD as "Sales" (Revenue realized immediately)
            sales_by_date[d] = sales_by_date.get(d, 0) + amount

    # Generate continuous date range
    labels = []
    sales_data =[]
    charges_data =[]
    
    current_date = start_time.date()
    end_date = now.date()
    
    while current_date <= end_date:
        labels.append(current_date.strftime(date_format))
        sales_data.append(sales_by_date.get(current_date, 0))
        charges_data.append(charges_by_date.get(current_date, 0))
        current_date += timedelta(days=1)
    
    return JsonResponse({
        'labels': labels, 
        'sales_data': sales_data,
        'charges_data': charges_data
    })

# --- MISSING UTILITY VIEWS ---

@require_POST
@permission_required('inventory.change_product', raise_exception=True)
def product_toggle_status(request, slug):
    product = get_object_or_404(Product, slug=slug)
    if product.status == Product.Status.ACTIVE:
        product.status = Product.Status.DEACTIVATED
        messages.success(request, f"'{product.name}' has been deactivated.")
    else:
        product.status = Product.Status.ACTIVE
        messages.success(request, f"'{product.name}' has been activated.")
    product.save()
    return redirect(product.get_absolute_url())

def process_history_records(history_records):
    """Helper to calculate deltas and action labels for history records."""
    # Eagerly load the records to work with a list
    records_on_page = list(history_records)
    if not records_on_page:
        return

    # --- OPTIMIZATION: Pre-fetch all previous records in bulk ---
    # 1. Get all unique product IDs from the current page of history
    product_ids = {r.id for r in records_on_page}
    
    # 2. Fetch all historical records for these products
    HistoryModel = records_on_page[0].__class__
    all_history_for_products = HistoryModel.objects.filter(
        id__in=product_ids
    ).order_by('id', 'history_date') # Order is crucial

    # 3. Create a map of {history_id: previous_record_object}
    prev_record_map = {}
    last_record_for_product = {}
    for record in all_history_for_products:
        product_id = record.id
        if product_id in last_record_for_product:
            # The current record's predecessor is the last one we saw for this product
            prev_record_map[record.history_id] = last_record_for_product[product_id]
        # Store the current record as the "last seen" for the next iteration
        last_record_for_product[product_id] = record

    # 4. Fetch Category names in bulk to resolve IDs in the change summary
    category_ids = set()
    for record in records_on_page:
        if record.category_id:
            category_ids.add(record.category_id)
        prev = prev_record_map.get(record.history_id)
        if prev and prev.category_id:
            category_ids.add(prev.category_id)
    categories_map = Category.objects.filter(id__in=category_ids).in_bulk()
    # --- END OPTIMIZATION ---

    for record in records_on_page:
        record.change_summary_html = "No details available."
        record.action_label = "Update"
        record.badge_class = "bg-secondary-subtle text-secondary border border-secondary"

        if record.history_type == '+':
            record.action_label = "Created"
            record.badge_class = "bg-success-subtle text-success border border-success"
            record.change_summary_html = "Initial product creation."
        elif record.history_type == '-':
            record.action_label = "Deleted"
            record.badge_class = "bg-danger-subtle text-danger border border-danger"
            record.change_summary_html = "Product deleted."
        elif record.history_type == '~':
            # Use the pre-fetched previous record instead of hitting the DB again
            prev_record = prev_record_map.get(record.history_id)
            
            if prev_record:
                delta = record.diff_against(prev_record)
                changes =[]
                affected_fields =[]
                
                for change in delta.changes:
                    field = change.field
                    if field in ['slug', 'date_updated']:
                        continue
                    
                    old_val = change.old
                    new_val = change.new
                    
                    if field == 'category':
                        old_cat = categories_map.get(old_val)
                        new_cat = categories_map.get(new_val)
                        old_val = old_cat.name if old_cat else "None"
                        new_val = new_cat.name if new_cat else "None"

                    changes.append(f"<strong>{field.replace('_', ' ').title()}:</strong> {old_val} &rarr; {new_val}")
                    
                    if field == 'price': affected_fields.append("Price")
                    elif field == 'quantity': affected_fields.append("Stock")
                    elif field == 'status': affected_fields.append("Status")
                    elif field == 'category': affected_fields.append("Category")
                    else: affected_fields.append("Details")
                
                record.change_summary_html = "<br>".join(changes) if changes else "No specific field changes detected."
                
                unique_fields = list(set(affected_fields))
                if not unique_fields:
                    record.action_label = "Update"
                elif len(unique_fields) == 1:
                    record.action_label = unique_fields[0]
                else:
                    record.action_label = "Multiple"
                
                # Badge Colors
                if "Price" in unique_fields: record.badge_class = "bg-info-subtle text-info-emphasis border border-info"
                elif "Stock" in unique_fields: record.badge_class = "bg-warning-subtle text-warning-emphasis border border-warning"
                elif "Status" in unique_fields: record.badge_class = "bg-dark-subtle text-dark-emphasis border border-dark"
                elif "Category" in unique_fields: record.badge_class = "bg-primary-subtle text-primary border border-primary"
            else:
                record.change_summary_html = "No previous record for comparison."

class TransactionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = StockTransaction
    template_name = 'inventory/transaction_list.html'
    context_object_name = 'transaction_list'
    paginate_by = 25
    permission_required = 'inventory.view_stocktransaction'
    
    def get_queryset(self):
        queryset = StockTransaction.objects.select_related('product', 'user').all()
        form = TransactionFilterForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data.get('product'): queryset = queryset.filter(product=form.cleaned_data['product'])
            if form.cleaned_data.get('transaction_type'): queryset = queryset.filter(transaction_type=form.cleaned_data['transaction_type'])
            if form.cleaned_data.get('transaction_reason'): queryset = queryset.filter(transaction_reason=form.cleaned_data['transaction_reason'])
            if form.cleaned_data.get('user'): queryset = queryset.filter(user=form.cleaned_data['user'])
            if form.cleaned_data.get('start_date'): queryset = queryset.filter(timestamp__date__gte=form.cleaned_data['start_date'])
            if form.cleaned_data.get('end_date'): queryset = queryset.filter(timestamp__date__lte=form.cleaned_data['end_date'])
        return queryset.order_by('-timestamp', '-id')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = TransactionFilterForm(self.request.GET)
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

class ProductHistoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Product.history.model
    template_name = 'inventory/product_history_list.html'
    context_object_name = 'history_list'
    paginate_by = 20
    permission_required = 'inventory.can_view_history'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('history_user')
        queryset = queryset.order_by('-history_date')
        form = ProductHistoryFilterForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data.get('product'):
                queryset = queryset.filter(id=form.cleaned_data['product'].id)
            if form.cleaned_data.get('user'):
                queryset = queryset.filter(history_user=form.cleaned_data['user'])
            if form.cleaned_data.get('start_date'):
                queryset = queryset.filter(history_date__date__gte=form.cleaned_data['start_date'])
            if form.cleaned_data.get('end_date'):
                queryset = queryset.filter(history_date__date__lte=form.cleaned_data['end_date'])
            
            # Handle Action Filtering
            action = form.cleaned_data.get('action')
            if action:
                queryset = queryset.filter(history_type=action)

        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProductHistoryFilterForm(self.request.GET)
        process_history_records(context['page_obj'])
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            query_params.pop('page')
        context['query_params'] = query_params.urlencode()
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context

class ProductHistoryDetailView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Product.history.model
    template_name = 'inventory/product_history_detail.html'
    context_object_name = 'history_list'
    paginate_by = 20
    permission_required = 'inventory.can_view_history'
    
    def dispatch(self, request, *args, **kwargs):
        self.product = get_object_or_404(Product, slug=self.kwargs['slug'])
        return super().dispatch(request, *args, **kwargs)
        
    def get_queryset(self):
        return self.product.history.select_related('history_user').all().order_by('-history_date')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = self.product
        process_history_records(context['page_obj'])
        if context.get('page_obj'):
            context['elided_page_range'] = context['page_obj'].paginator.get_elided_page_range(context['page_obj'].number, on_each_side=1, on_ends=1)
        return context